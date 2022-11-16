from openpyxl import load_workbook
import re

workbook = load_workbook(filename="data.xlsx")
sheet = workbook.active

current_row = 2

for value in sheet.iter_rows(min_row = 2, max_row = 1414, min_col = 5, max_col = 5):
    text = value[0].value

    # text = text.replace("Gift Article Share" + "\n" + "\n", "")
    # text = text.replace("GiftOutline Gift Article", "")
    # text = text.replace("Comment on this story Comment", "")
    # text = text.replace("Advertisement" + "\n" + "\n", "")
    # text = text.replace("To hear more audio stories from publishers like The New York Times, download Audm for iPhone or Android." + "\n" + "\n", "")
    # text = text.replace("Warning: This graphic requires JavaScript. Please enable JavaScript for the best experience." + "\n" + "\n", "")
    # for num in range(21):
    #     text = text.replace("Listen " + str(num) + " min  ", "")
    text = text.replace("Story continues below advertisement" + "\n" + "\n", "")


    sheet["E" + str(current_row)] = text
    print("E" + str(current_row))

    workbook.save(filename="data.xlsx")

    current_row += 1

