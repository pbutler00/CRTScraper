from openpyxl import load_workbook
import re
from htmldate import find_date

workbook = load_workbook(filename="data.xlsx")
sheet = workbook.active

current_row = 1231
#1096

for value in sheet.iter_rows(min_row = 1231, max_row = 1235, min_col = 4, max_col = 4):
    date = value[0].value

    if date == "None" and sheet["I" + str(current_row)].value == "www.foxnews.com":
        sheet["D" + str(current_row)] = find_date(sheet["H" + str(current_row)].value)

    print("D" + str(current_row))

    workbook.save(filename="data.xlsx")

    current_row += 1
